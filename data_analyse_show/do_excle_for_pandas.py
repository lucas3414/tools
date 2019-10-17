import pandas as pd
from openpyxl import load_workbook


class ReadExcle:

    def __init__(self, filename, sheet):
        self.filename = filename
        self.sheet = sheet

    def to_data_frame(self):
        try:
            self.sheet in pd.ExcelFile(self.filename).sheet_names
            return pd.DataFrame(pd.read_excel(self.filename, self.sheet))
        except Exception as e:
            print(e)
            return None

    def read_date(self, *args):
        test_data = []
        if len(args) == 0:
            args = [column for column in self.to_data_frame()]
        for i in range(int(self.to_data_frame().shape[0])):
            row_data = self.to_data_frame().loc[i, args].to_dict()
            test_data.append(row_data)
        return test_data

    def updata_date(self, columns, num, response_data):
        # 读取原来数据
        old_data = pd.DataFrame(self.read_date())
        # 修改数据
        old_data.loc[int(num), columns] = response_data
        # pd 将数据写入excle
        work_book = pd.ExcelWriter(self.filename, engine='openpyxl')
        # 以 openpyxl 方式打开excle
        book = load_workbook(work_book.path)
        work_book.book = book
        # 删除同名sheet 页
        del book[self.sheet]
        # 写入新数据组成的sheet页
        old_data.to_excel(excel_writer=work_book, sheet_name=self.sheet, index=None)
        work_book.close()

    def create_new_sheet(self, columns, response_data):
        data = pd.DataFrame({columns: response_data}, index=[0])
        work_book = pd.ExcelWriter(self.filename, engine='openpyxl')
        book = load_workbook(work_book.path)
        work_book.book = book
        data.to_excel(excel_writer=work_book, sheet_name=self.sheet, index=None)
        work_book.close()



if __name__ == '__main__':
    # RE = ReadExcle('demo.xlsx', 'index')
    # print(RE.read_date())

    # RE.updata_date('kk', 2, 'kkkdasdasdasdada')
    # RE.updata_date('哈哈', 2, '一下')
    # RE.updata_date('哈哈', 3, '一下的撒打算')
    # RE.updata_date('token', 2, '3414')
    #
    RE = ReadExcle('demo.xlsx', 'dddddadd')
    RE.create_new_sheet('ddd', 'aasdadsad')
    pass