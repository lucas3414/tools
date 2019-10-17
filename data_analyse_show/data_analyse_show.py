import pandas as pd
from openpyxl import load_workbook
from pyecharts.charts import Bar
from pyecharts import options as opts
import collections


class DoExcle:

    def __init__(self, filename, sheet):
        self.filename = filename
        self.sheet = sheet

    def to_data_frame(self):
        try:
            self.sheet in pd.ExcelFile(self.filename).sheet_names
            return pd.DataFrame(pd.read_excel(self.filename, self.sheet))
        except Exception as e:
            print(e)
            return None

    def read_date(self, *args):
        test_data = []
        if len(args) == 0:
            args = [column for column in self.to_data_frame()]
        for i in range(int(self.to_data_frame().shape[0])):
            row_data = self.to_data_frame().loc[i, args].to_dict()
            test_data.append(row_data)
        return test_data

    def updata_date(self, columns, num, response_data):
        old_data = pd.DataFrame(self.read_date())
        old_data.loc[int(num), columns] = response_data
        work_book = pd.ExcelWriter(self.filename, engine='openpyxl')
        book = load_workbook(work_book.path)
        work_book.book = book
        del book[self.sheet]
        old_data.to_excel(excel_writer=work_book, sheet_name=self.sheet, index=None)
        work_book.close()

    def create_new_sheet(self, columns, response_data):
        data = pd.DataFrame({columns: response_data}, index=[0])
        work_book = pd.ExcelWriter(self.filename, engine='openpyxl')
        book = load_workbook(work_book.path)
        work_book.book = book
        data.to_excel(excel_writer=work_book, sheet_name=self.sheet, index=None)
        work_book.close()


# 数据直方图展示
def show_data_analyse_result_of_bar(data_list, xdata, ydata):
    '''
    这个是用直方图展示数据
    :param data_list:  数据源，list:
        [{'type': '信息类', 'module': '数据监控台', 'id': 228, 'state': 'opened', 'created': '2018-08-21T02:33:03.395Z',
        'updated': '2018-08-21T02:34:29.101Z', 'labels1': '优先级-普通', 'labels2': '状态-打开', 'labels3': '类别-BUG',
        'milestone': 'v1.2.6', 'assignee': 'liuyang', 'notes': 0},]
    :param xdata:  直方图的横坐标
    :param ydata:  直方图的纵坐标
    '''
    info = {}
    bar = Bar()
    for ele in list(set([ele[ydata] for ele in data_list])):
        info[ele] = dict(collections.Counter(
            [ele[xdata] for ele in filter(lambda x: x[ydata] == ele, data_list)]
        ))
    bar.add_xaxis(list(set([ele[xdata] for ele in data_list])))
    for module, module_info in info.items():
        module_data = []
        for ele in list(set([ele[xdata] for ele in data_list])):
            module_data.append(module_info.get(ele, None))
        bar.add_yaxis(module, module_data)
    bar.set_global_opts(
        # title_opts=opts.TitleOpts(title="版本统计", subtitle="副标题", ),
        toolbox_opts=opts.ToolboxOpts(is_show=False, ),
        legend_opts=opts.LegendOpts(is_show=True, )
    )
    bar.render()


if __name__ == '__main__':
    RE = DoExcle('case.xlsx', 'case')
    # <class 'list'>
    data_list = RE.read_date()

    show_data_analyse_result_of_bar(data_list=data_list, xdata='milestone', ydata='module')
